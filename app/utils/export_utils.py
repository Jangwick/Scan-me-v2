"""
Export Utilities for ScanMe System
Handles data export to various formats (Excel, CSV, PDF)
"""

import pandas as pd
from datetime import datetime, date
import os
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Get the base directory for exports
def get_export_dir():
    """Get the absolute path to the exports directory"""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    export_dir = os.path.join(base_dir, 'exports')
    os.makedirs(export_dir, exist_ok=True)
    return export_dir

def export_attendance_to_excel(attendance_data, filename=None):
    """
    Export attendance data to Excel file
    Args:
        attendance_data (list): List of attendance record dictionaries
        filename (str): Output filename
    Returns:
        str: Path to created file
    """
    try:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"attendance_report_{timestamp}.xlsx"
        
        # Ensure export directory exists
        export_dir = get_export_dir()
        output_path = os.path.join(export_dir, filename)
        
        # Convert to DataFrame
        df = pd.DataFrame(attendance_data)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Add header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Define styles for different attendance statuses
        absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Light red
        absent_font = Font(color="991B1B", bold=True)  # Dark red
        
        late_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Light yellow
        late_font = Font(color="92400E")  # Dark yellow
        
        present_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Light green
        present_font = Font(color="065F46")  # Dark green
        
        # Write data
        for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(r)
            
            # Apply conditional formatting based on attendance status
            if r_idx > 1:  # Skip header row
                # Find the Attendance Status column
                try:
                    status_col_idx = list(df.columns).index('Attendance Status') + 1 if 'Attendance Status' in df.columns else None
                    
                    if status_col_idx:
                        status_cell = ws.cell(row=r_idx, column=status_col_idx)
                        status_value = str(status_cell.value) if status_cell.value else ""
                        
                        # Apply formatting to entire row based on status
                        for col_idx in range(1, len(df.columns) + 1):
                            cell = ws.cell(row=r_idx, column=col_idx)
                            
                            if 'Absent' in status_value:
                                cell.fill = absent_fill
                                if col_idx == status_col_idx:
                                    cell.font = absent_font
                            elif 'Late' in status_value:
                                cell.fill = late_fill
                                if col_idx == status_col_idx:
                                    cell.font = late_font
                            elif 'Present' in status_value and 'On-Time' in status_value:
                                cell.fill = present_fill
                                if col_idx == status_col_idx:
                                    cell.font = present_font
                except (ValueError, IndexError):
                    pass  # Column not found, skip formatting
        
        # Format header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        wb.save(output_path)
        
        return output_path
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

def export_attendance_to_csv(attendance_data, filename=None):
    """
    Export attendance data to CSV file
    Args:
        attendance_data (list): List of attendance record dictionaries
        filename (str): Output filename
    Returns:
        str: Path to created file
    """
    try:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"attendance_report_{timestamp}.csv"
        
        # Ensure export directory exists
        export_dir = get_export_dir()
        output_path = os.path.join(export_dir, filename)
        
        if not attendance_data:
            return None
        
        # Write CSV file
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            if attendance_data:
                fieldnames = attendance_data[0].keys()
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(attendance_data)
        
        return output_path
        
    except Exception as e:
        print(f"Error exporting to CSV: {e}")
        import traceback
        traceback.print_exc()
        return None

def export_attendance_to_pdf(attendance_data, title="Attendance Report", filename=None):
    """
    Export attendance data to PDF file
    Args:
        attendance_data (list): List of attendance record dictionaries
        title (str): Report title
        filename (str): Output filename
    Returns:
        str: Path to created file
    """
    try:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"attendance_report_{timestamp}.pdf"
        
        # Ensure export directory exists
        export_dir = get_export_dir()
        output_path = os.path.join(export_dir, filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        elements = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add title
        title_para = Paragraph(title, title_style)
        elements.append(title_para)
        
        # Add generation date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            spaceAfter=20
        )
        date_para = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style)
        elements.append(date_para)
        
        if attendance_data:
            # Calculate statistics
            total_records = len(attendance_data)
            absent_count = len([r for r in attendance_data if 'Absent' in str(r.get('Attendance Status', ''))])
            present_count = total_records - absent_count
            
            # Add summary
            summary_text = f"<b>Summary:</b> Total Records: {total_records} | Present: {present_count} | Absent (No Time-Out): {absent_count}"
            summary_para = Paragraph(summary_text, styles['Normal'])
            elements.append(summary_para)
            elements.append(Spacer(1, 20))
            
            # Prepare table data
            headers = list(attendance_data[0].keys())
            table_data = [headers]
            
            # Find Attendance Status column index
            status_col_idx = headers.index('Attendance Status') if 'Attendance Status' in headers else None
            
            for record in attendance_data:
                row = [str(record.get(header, '')) for header in headers]
                table_data.append(row)
            
            # Create table
            table = Table(table_data)
            
            # Base table style
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]
            
            # Add conditional row coloring based on attendance status
            if status_col_idx is not None:
                for i, record in enumerate(attendance_data, start=1):
                    status = str(record.get('Attendance Status', ''))
                    
                    if 'Absent' in status:
                        # Light red background for absent students
                        table_style.append(('BACKGROUND', (0, i), (-1, i), colors.Color(1, 0.89, 0.89)))
                        table_style.append(('TEXTCOLOR', (status_col_idx, i), (status_col_idx, i), colors.Color(0.6, 0.1, 0.1)))
                    elif 'Late' in status:
                        # Light yellow background for late students
                        table_style.append(('BACKGROUND', (0, i), (-1, i), colors.Color(1, 0.95, 0.78)))
                    elif 'On-Time' in status:
                        # Light green background for on-time students
                        table_style.append(('BACKGROUND', (0, i), (-1, i), colors.Color(0.82, 0.98, 0.9)))
            
            table.setStyle(TableStyle(table_style))
            
            elements.append(table)
        else:
            no_data_para = Paragraph("No attendance data found.", styles['Normal'])
            elements.append(no_data_para)
        
        # Build PDF
        doc.build(elements)
        
        return output_path
        
    except Exception as e:
        print(f"Error exporting to PDF: {e}")
        import traceback
        traceback.print_exc()
        return None

def generate_student_report(student, attendance_records, start_date=None, end_date=None):
    """
    Generate individual student attendance report
    Args:
        student: Student object
        attendance_records: List of attendance records
        start_date: Start date for report
        end_date: End date for report
    Returns:
        dict: Report data
    """
    try:
        # Filter records by date range if provided
        if start_date:
            attendance_records = [r for r in attendance_records if r.scan_time.date() >= start_date]
        if end_date:
            attendance_records = [r for r in attendance_records if r.scan_time.date() <= end_date]
        
        # Calculate statistics
        total_days = len(set(r.scan_time.date() for r in attendance_records))
        total_scans = len(attendance_records)
        late_arrivals = len([r for r in attendance_records if r.is_late])
        absent_records = len([r for r in attendance_records if r.time_out is None])
        complete_attendance = total_scans - absent_records
        rooms_visited = len(set(r.room_id for r in attendance_records))
        
        # Group by date
        daily_attendance = {}
        for record in attendance_records:
            date_key = record.scan_time.date()
            if date_key not in daily_attendance:
                daily_attendance[date_key] = []
            daily_attendance[date_key].append(record)
        
        report_data = {
            'student_info': {
                'name': student.get_full_name(),
                'student_no': student.student_no,
                'department': student.department,
                'section': student.section,
                'year_level': student.year_level
            },
            'period': {
                'start_date': start_date.isoformat() if start_date else None,
                'end_date': end_date.isoformat() if end_date else None
            },
            'summary': {
                'total_days_attended': total_days,
                'total_scans': total_scans,
                'complete_attendance': complete_attendance,
                'incomplete_attendance': absent_records,
                'late_arrivals': late_arrivals,
                'on_time_percentage': round(((total_scans - late_arrivals) / total_scans * 100), 2) if total_scans > 0 else 0,
                'completion_rate': round((complete_attendance / total_scans * 100), 2) if total_scans > 0 else 0,
                'rooms_visited': rooms_visited
            },
            'daily_breakdown': {
                str(date): {
                    'scans': len(records),
                    'rooms': [r.room.get_full_name() for r in records],
                    'times': [r.scan_time.strftime('%H:%M:%S') for r in records],
                    'late_count': len([r for r in records if r.is_late]),
                    'incomplete_count': len([r for r in records if r.time_out is None])
                }
                for date, records in daily_attendance.items()
            },
            'records': [
                {
                    'date': r.scan_time.date().isoformat(),
                    'time_in': r.time_in.time().isoformat() if r.time_in else 'N/A',
                    'time_out': r.time_out.time().isoformat() if r.time_out else 'No Time-Out',
                    'duration': r.get_duration() if r.time_out else 0,
                    'room': r.room.get_full_name() if r.room else 'Unknown',
                    'is_late': r.is_late,
                    'status': 'Absent (No Time-Out)' if r.time_out is None else ('Present (Late)' if r.is_late else 'Present (On-Time)'),
                    'scanner': r.scanned_by_user.username if r.scanned_by_user else 'System'
                }
                for r in sorted(attendance_records, key=lambda x: x.scan_time, reverse=True)
            ]
        }
        
        return report_data
        
    except Exception as e:
        print(f"Error generating student report: {e}")
        return None

def generate_room_report(room, attendance_records, start_date=None, end_date=None):
    """
    Generate room attendance report
    Args:
        room: Room object
        attendance_records: List of attendance records
        start_date: Start date for report
        end_date: End date for report
    Returns:
        dict: Report data
    """
    try:
        # Filter records by date range if provided
        if start_date:
            attendance_records = [r for r in attendance_records if r.scan_time.date() >= start_date]
        if end_date:
            attendance_records = [r for r in attendance_records if r.scan_time.date() <= end_date]
        
        # Calculate statistics
        total_scans = len(attendance_records)
        unique_students = len(set(r.student_id for r in attendance_records))
        unique_days = len(set(r.scan_time.date() for r in attendance_records))
        late_arrivals = len([r for r in attendance_records if r.is_late])
        incomplete_attendance = len([r for r in attendance_records if r.time_out is None])
        complete_attendance = total_scans - incomplete_attendance
        
        # Group by date
        daily_stats = {}
        for record in attendance_records:
            date_key = record.scan_time.date()
            if date_key not in daily_stats:
                daily_stats[date_key] = {
                    'students': set(),
                    'total_scans': 0,
                    'late_count': 0,
                    'incomplete_count': 0
                }
            
            daily_stats[date_key]['students'].add(record.student_id)
            daily_stats[date_key]['total_scans'] += 1
            if record.is_late:
                daily_stats[date_key]['late_count'] += 1
            if record.time_out is None:
                daily_stats[date_key]['incomplete_count'] += 1
        
        # Convert sets to counts
        for date, stats in daily_stats.items():
            stats['unique_students'] = len(stats['students'])
            del stats['students']
        
        report_data = {
            'room_info': {
                'name': room.get_full_name(),
                'room_number': room.room_number,
                'building': room.building,
                'capacity': room.capacity,
                'room_type': room.room_type
            },
            'period': {
                'start_date': start_date.isoformat() if start_date else None,
                'end_date': end_date.isoformat() if end_date else None
            },
            'summary': {
                'total_scans': total_scans,
                'unique_students': unique_students,
                'unique_days': unique_days,
                'late_arrivals': late_arrivals,
                'complete_attendance': complete_attendance,
                'incomplete_attendance': incomplete_attendance,
                'completion_rate': round((complete_attendance / total_scans * 100), 2) if total_scans > 0 else 0,
                'average_daily_attendance': round(unique_students / unique_days, 2) if unique_days > 0 else 0,
                'capacity_utilization': round((unique_students / room.capacity * 100), 2) if room.capacity > 0 else 0
            },
            'daily_breakdown': {
                str(date): stats for date, stats in daily_stats.items()
            },
            'records': [
                {
                    'date': r.scan_time.date().isoformat(),
                    'time_in': r.time_in.time().isoformat() if r.time_in else 'N/A',
                    'time_out': r.time_out.time().isoformat() if r.time_out else 'No Time-Out',
                    'duration': r.get_duration() if r.time_out else 0,
                    'student': r.student.get_full_name() if r.student else 'Unknown',
                    'student_no': r.student.student_no if r.student else 'N/A',
                    'is_late': r.is_late,
                    'status': 'Absent (No Time-Out)' if r.time_out is None else ('Present (Late)' if r.is_late else 'Present (On-Time)'),
                    'scanner': r.scanned_by_user.username if r.scanned_by_user else 'System'
                }
                for r in sorted(attendance_records, key=lambda x: x.scan_time, reverse=True)
            ]
        }
        
        return report_data
        
    except Exception as e:
        print(f"Error generating room report: {e}")
        return None

def export_students_to_excel(students_data, filename=None):
    """
    Export students data to Excel file
    Args:
        students_data (list): List of student dictionaries
        filename (str): Output filename
    Returns:
        str: Path to created file
    """
    try:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"students_list_{timestamp}.xlsx"
        
        output_path = os.path.join('exports', filename)
        os.makedirs('exports', exist_ok=True)
        
        # Convert to DataFrame
        df = pd.DataFrame(students_data)
        
        # Write to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Students')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Students']
            
            # Format header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return output_path
        
    except Exception as e:
        print(f"Error exporting students to Excel: {e}")
        return None

def get_export_summary(export_path, record_count):
    """
    Get export summary information
    Args:
        export_path (str): Path to exported file
        record_count (int): Number of records exported
    Returns:
        dict: Export summary
    """
    try:
        file_stats = os.stat(export_path)
        
        return {
            'file_path': export_path,
            'file_name': os.path.basename(export_path),
            'file_size': file_stats.st_size,
            'file_size_mb': round(file_stats.st_size / (1024 * 1024), 2),
            'record_count': record_count,
            'created_at': datetime.fromtimestamp(file_stats.st_ctime).isoformat(),
            'success': True
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }